from typing import Any

from openpyxl import load_workbook


class ParseSpreadsheet:
    def __init__(
        self,
        spreadsheet_path: str = r"./spreadsheet/NFL_Poop_Picks_2025-26.xlsx",
        sheet: str = "Week 1",
    ):
        self.path = spreadsheet_path
        self.workbook = load_workbook(spreadsheet_path)
        self.sheet = self.workbook[sheet]

        self.PICK_COLUMNS = [4, 5, 6, 7]  # E, F, G, H
        self.RESULT_COLUMNS = [9, 10, 11, 12]  # J, K, L, M

        # TODO: Stop accessing values via letter reference, use num index instead

    def get_column(
        self, col: int, start_row: int = 1, end_row: int | None = None
    ) -> list:
        """Get all values in a given column"""
        if end_row is None:
            end_row = self.sheet.max_row

        return [
            self.sheet.cell(row, col).value for row in range(start_row, end_row + 1)
        ]

    def get_cell(self, cell_ref: str) -> str:
        """Get value stored in given cell.

        Args
            `cell_ref`: The cell to get the value of, must be in Excel format
                Example: 'B2', 'C5', ..."""
        return self.sheet[cell_ref].value

    def set_cell(self, cell_ref: str, value: Any) -> None:
        self.sheet[cell_ref] = value


ps = ParseSpreadsheet()
